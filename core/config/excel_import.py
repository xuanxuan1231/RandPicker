"""
Excel 导入导出功能
"""

from pathlib import Path
from typing import List, Dict, Any
from uuid import uuid4

from loguru import logger

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    logger.warning("openpyxl 未安装，Excel 导入导出功能将不可用。")
    OPENPYXL_AVAILABLE = False


class ExcelImporter:
    """Excel 导入导出工具类"""

    @staticmethod
    def import_from_excel(file_path: str, import_mode: str = "replace") -> tuple[List[Dict[str, Any]], str]:
        """
        从 Excel 文件导入学生列表。

        :param file_path: Excel 文件路径
        :param import_mode: 导入模式 "replace" (替换全部) 或 "merge" (合并追加)
        :return: (学生列表, 错误信息) 如果成功则错误信息为空字符串
        """
        if not OPENPYXL_AVAILABLE:
            return [], "openpyxl 库未安装，无法导入 Excel 文件。"

        try:
            file = Path(file_path)
            if not file.exists():
                return [], f"文件不存在: {file_path}"

            if not file.suffix.lower() in ['.xlsx', '.xls']:
                return [], "不支持的文件格式。请选择 .xlsx 或 .xls 文件。"

            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active

            # 检查是否为空工作表
            if sheet.max_row < 2:
                return [], "Excel 文件为空或只有标题行。"

            # 读取标题行（第一行）
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value).strip() if cell.value else "")

            # 查找必需的列
            name_col = -1
            weight_col = -1
            enabled_col = -1
            avatar_col = -1
            properties_start_col = -1

            for idx, header in enumerate(headers):
                header_lower = header.lower()
                if header_lower in ['姓名', 'name', '名字', '学生姓名']:
                    name_col = idx
                elif header_lower in ['权重', 'weight']:
                    weight_col = idx
                elif header_lower in ['启用', 'enabled', '是否启用']:
                    enabled_col = idx
                elif header_lower in ['头像', 'avatar', '头像路径', 'avatar_path']:
                    avatar_col = idx
                elif header_lower.startswith('属性') or header_lower.startswith('property'):
                    if properties_start_col == -1:
                        properties_start_col = idx

            # 姓名列是必需的
            if name_col == -1:
                return [], "Excel 文件缺少必需的「姓名」列。请确保第一行包含「姓名」或「name」列。"

            # 读取数据行
            students = []
            error_rows = []

            for row_idx in range(2, sheet.max_row + 1):
                row = sheet[row_idx]

                # 读取姓名
                name = str(row[name_col].value).strip() if row[name_col].value else ""
                if not name:
                    error_rows.append(f"第 {row_idx} 行：姓名为空，已跳过")
                    continue

                # 读取权重
                weight = 1.0
                if weight_col != -1 and row[weight_col].value is not None:
                    try:
                        weight = float(row[weight_col].value)
                        if weight < 0:
                            weight = 0.0
                    except (ValueError, TypeError):
                        error_rows.append(f"第 {row_idx} 行：权重值无效，已使用默认值 1.0")
                        weight = 1.0

                # 读取启用状态
                enabled = True
                if enabled_col != -1 and row[enabled_col].value is not None:
                    val = str(row[enabled_col].value).strip().upper()
                    if val in ['FALSE', '0', 'NO', 'N', '否', '禁用']:
                        enabled = False
                    elif val in ['TRUE', '1', 'YES', 'Y', '是', '启用']:
                        enabled = True

                # 读取头像路径
                avatar = ""
                if avatar_col != -1 and row[avatar_col].value is not None:
                    avatar = str(row[avatar_col].value).strip()

                # 读取附加属性（从属性列开始的所有列，成对读取）
                properties = []
                if properties_start_col != -1:
                    for prop_idx in range(properties_start_col, len(row), 2):
                        if prop_idx + 1 < len(row):
                            prop_name = str(row[prop_idx].value).strip() if row[prop_idx].value else ""
                            prop_value = str(row[prop_idx + 1].value).strip() if row[prop_idx + 1].value else ""
                            if prop_name and prop_value:
                                properties.append({
                                    "name": prop_name,
                                    "value": prop_value
                                })

                # 创建学生对象
                student = {
                    "id": str(uuid4()),
                    "name": name,
                    "weight": weight,
                    "enabled": enabled,
                    "avatar": avatar,
                    "properties": properties
                }
                students.append(student)

            workbook.close()

            if not students:
                return [], "未能从 Excel 文件中读取到有效的学生数据。"

            # 构建返回信息
            info_msg = f"成功导入 {len(students)} 名学生。"
            if error_rows:
                info_msg += f"\n\n注意事项：\n" + "\n".join(error_rows[:5])
                if len(error_rows) > 5:
                    info_msg += f"\n... 还有 {len(error_rows) - 5} 条警告"

            logger.success(f"从 Excel 导入了 {len(students)} 名学生")
            return students, info_msg

        except Exception as e:
            logger.exception(f"导入 Excel 文件时发生错误: {e}")
            return [], f"导入失败: {str(e)}"

    @staticmethod
    def export_to_excel(file_path: str, students: List[Dict[str, Any]]) -> tuple[bool, str]:
        """
        导出学生列表到 Excel 文件。

        :param file_path: 输出文件路径
        :param students: 学生列表
        :return: (是否成功, 消息)
        """
        if not OPENPYXL_AVAILABLE:
            return False, "openpyxl 库未安装，无法导出 Excel 文件。"

        try:
            if not students:
                return False, "没有学生数据可导出。"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "学生列表"

            # 写入标题行
            headers = ["姓名", "权重", "启用", "头像路径"]

            # 查找最大属性数量
            max_properties = 0
            for student in students:
                props = student.get("properties", [])
                if len(props) > max_properties:
                    max_properties = len(props)

            # 为每个属性添加两列（名称和值）
            for i in range(max_properties):
                headers.append(f"属性{i + 1}名称")
                headers.append(f"属性{i + 1}值")

            # 写入标题
            for col_idx, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_idx, value=header)

            # 写入数据行
            for row_idx, student in enumerate(students, start=2):
                sheet.cell(row=row_idx, column=1, value=student.get("name", ""))
                sheet.cell(row=row_idx, column=2, value=student.get("weight", 1.0))
                sheet.cell(row=row_idx, column=3, value="是" if student.get("enabled", True) else "否")
                sheet.cell(row=row_idx, column=4, value=student.get("avatar", ""))

                # 写入附加属性
                properties = student.get("properties", [])
                for prop_idx, prop in enumerate(properties):
                    col_offset = 5 + (prop_idx * 2)
                    sheet.cell(row=row_idx, column=col_offset, value=prop.get("name", ""))
                    sheet.cell(row=row_idx, column=col_offset + 1, value=prop.get("value", ""))

            # 自动调整列宽
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

            # 保存文件
            workbook.save(file_path)
            workbook.close()

            logger.success(f"成功导出 {len(students)} 名学生到 {file_path}")
            return True, f"成功导出 {len(students)} 名学生。"

        except Exception as e:
            logger.exception(f"导出 Excel 文件时发生错误: {e}")
            return False, f"导出失败: {str(e)}"

    @staticmethod
    def is_available() -> bool:
        """检查 Excel 导入导出功能是否可用"""
        return OPENPYXL_AVAILABLE

    @staticmethod
    def get_template_headers() -> List[str]:
        """返回 Excel 模板的标题行"""
        return ["姓名", "权重", "启用", "头像路径", "属性1名称", "属性1值"]
